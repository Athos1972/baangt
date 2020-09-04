import logging
from xlrd3 import open_workbook
import json
import baangt.base.CustGlobalConstants as CGC
import baangt.base.GlobalConstants as GC
from baangt.base.Utils import utils
import re
from random import randint
from openpyxl import load_workbook
from baangt.TestDataGenerator.TestDataGenerator import TestDataGenerator
from threading import Thread
import os

logger = logging.getLogger("pyC")


class Writer:
    """
    This class is made to update existing excel file.
    First it will open the file in python and then we can do multiple writes and once everything is update we can use
    save method in order to save the updated excel file. Hence, this class is very useful is saving time while updating
    excel files.
    """
    def __init__(self, path):
        self.path = path
        self.workbook = load_workbook(self.path)

    def write(self, row, data, sht):
        # Update the values using row and col number.
        # Note :- We are using openpyxl so row & column index will start from 1 instead of 0
        column = 0
        sheet = self.workbook[sht]
        headers = next(sheet.rows)
        for header in headers:  # finds the header position
            if "usecount" in str(header.value).lower():
                column = headers.index(header) + 1
        if column:
            sheet.cell(row, column).value = data

    def save(self):
        # Call this method to save the file once every updates are written
        self.workbook.save(self.path)
        self.workbook.close()


class HandleDatabase:
    def __init__(self, linesToRead, globalSettings=None):
        self.lineNumber = 3
        # FIXME: This is still not clean. GlobalSettings shouldn't be predefined in CustomConstants-Class
        self.globals = {
            CGC.CUST_TOASTS: "",
            GC.EXECUTION_STAGE: "",
            GC.TESTCASEERRORLOG: "",
            CGC.VIGOGFNUMMER: "",
            CGC.SAPPOLNR: "",
            CGC.PRAEMIE: "",
            CGC.POLNRHOST: "",
            GC.TESTCASESTATUS: "",
            GC.TIMING_DURATION: "",
            GC.SCREENSHOTS: "",
            GC.TIMELOG: "",
        }
        if globalSettings:
            for setting, value in globalSettings.items():
                self.globals[setting] = value
        self.range = self.__buildRangeOfRecords(linesToRead)
        self.rangeDict = {}
        self.__buildRangeDict()
        self.df_json = None
        self.dataDict = []
        self.recordPointer = 0
        self.sheet_dict = {}
        self.usecount = False

    def __buildRangeDict(self):
        """
        Interprets the Range and creates a DICT of values, that we can loop over later

        @return: Creates empty self.rangeDict
        """
        for lRangeLine in self.range:
            for x in range(lRangeLine[0], lRangeLine[1]+1):
                self.rangeDict[x] = ""

    def __buildRangeOfRecords(self, rangeFromConfigFile):
        lRange = []
        if not rangeFromConfigFile:
            # No selection - means all records
            return [[0,99999]]
        else:
            # Format: 4;6-99;17-200;203 or
            # Format: 4,6-100,800-1000
            if ";" in rangeFromConfigFile:
                for kombination in rangeFromConfigFile.split(";"):
                    lRange.append(HandleDatabase.__buildRangeOfRecordsOneEntry(kombination))
            elif "," in rangeFromConfigFile:
                for kombination in rangeFromConfigFile.split(","):
                    lRange.append(HandleDatabase.__buildRangeOfRecordsOneEntry(kombination))
            else:
                lRange.append(HandleDatabase.__buildRangeOfRecordsOneEntry(rangeFromConfigFile))

        # Make sure these are numbers:
        for lRangeLine in lRange:
            lRangeLine[0] = HandleDatabase.__sanitizeNumbers(lRangeLine[0])
            lRangeLine[1] = HandleDatabase.__sanitizeNumbers(lRangeLine[1])

        return lRange

    @staticmethod
    def __sanitizeNumbers(numberIn):
        if isinstance(numberIn, dict):
            numberIn = numberIn['default']
            try:
                  return int(numberIn.strip())
            except:
                  return 0
        numberIn = numberIn.strip()
        return int(numberIn)

    @staticmethod
    def __buildRangeOfRecordsOneEntry(rangeIn):
        if "-" in rangeIn:
            # This is a range (17-22)
            return HandleDatabase.__buildRangeOfRecordsSingleRange(rangeIn)
        else:
            # This is a single entry:
            return [rangeIn, rangeIn]

    @staticmethod
    def __buildRangeOfRecordsSingleRange(rangeIn):
        lSplit = rangeIn.split("-")
        return [lSplit[0], lSplit[1]]

    def read_excel(self, fileName, sheetName):
        fileName = utils.findFileAndPathFromPath(fileName)
        if not fileName:
            logger.critical(f"Can't open file: {fileName}")
            return
        logger.debug(f"Reading excel file {fileName}...")
        book = open_workbook(fileName)
        sheet = book.sheet_by_name(sheetName)

        # read header values into the list
        keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]

        # if testresult header is present then taking its index, which is later used as column number
        testrun_index = [keys.index(x) for x in keys if str(x).lower() == "testresult"]
        if testrun_index:
            testrun_index = testrun_index[0] + 1  # adding +1 value which is the correct column position
        else:  # if list is empty that means their is no testresult header
            testrun_index = 0

        for row_index in range(1, sheet.nrows):
            temp_dic = {}
            for col_index in range(sheet.ncols):
                temp_dic[keys[col_index]] = sheet.cell(row_index, col_index).value
                if type(temp_dic[keys[col_index]]) == float:
                    temp_dic[keys[col_index]] = repr(temp_dic[keys[col_index]])
                    if temp_dic[keys[col_index]][-2:] == ".0":
                        temp_dic[keys[col_index]] = temp_dic[keys[col_index]][:-2]
            # row, column, sheetName & fileName which are later used in updating source testrun file
            temp_dic["testcase_row"] = row_index
            temp_dic["testcase_sheet"] = sheetName
            temp_dic["testcase_file"] = fileName
            temp_dic["testcase_column"] = testrun_index
            self.dataDict.append(temp_dic)

    def update_datarecords(self, dataDict, fileName, sheetName):
        logger.debug("Updating prefix data...")
        testDataGenerator = TestDataGenerator(fileName, sheetName=sheetName, from_handleDatabase=True)
        for td in dataDict:
            temp_dic = dataDict[td]
            new_data_dic = {}
            for keys in temp_dic:
                if type(temp_dic[keys]) != str:
                    continue
                if '$(' in str(temp_dic[keys]):
                    while '$(' in str(temp_dic[keys]):
                        start_index = temp_dic[keys].index('$(')
                        end_index = temp_dic[keys][start_index:].index(')')+start_index
                        data_to_replace_with = temp_dic[temp_dic[keys][start_index+2:end_index]]
                        temp_dic[keys] = temp_dic[keys].replace(
                            temp_dic[keys][start_index:end_index+1], data_to_replace_with
                        )

                if str(temp_dic[keys])[:4].upper() == "RRD_":
                    logger.debug(f"Processing rrd data - {temp_dic[keys]}")
                    rrd_data = self.get_data_from_tdg(temp_dic[keys], testDataGenerator)
                    testDataGenerator.usecount_dict[repr(rrd_data)]["use"] += 1
                    testDataGenerator.update_usecount_in_source(rrd_data)
                    for data in rrd_data:
                        new_data_dic[data] = rrd_data[data]
                    logger.debug(f"Data processed - {temp_dic[keys]}")
                elif str(temp_dic[keys])[:4].upper() == "RRE_":
                    logger.debug(f"Processing rre data - {temp_dic[keys]}")
                    rre_data = self.get_data_from_tdg(temp_dic[keys], testDataGenerator)
                    testDataGenerator.usecount_dict[repr(rre_data)]["use"] += 1
                    testDataGenerator.update_usecount_in_source_rre(rre_data)
                    for data in rre_data:
                        new_data_dic[data] = rre_data[data]
                    logger.debug(f"Data processed - {temp_dic[keys]}")
                elif str(temp_dic[keys])[:4].upper() == "RLP_":
                    temp_dic[keys] = self.rlp_process(temp_dic[keys], fileName)
                elif str(temp_dic[keys])[:5].upper() == "RENV_":
                    temp_dic[keys] = str(TestDataGenerator.get_env_variable(temp_dic[keys][5:]))
                else:
                    try:
                        js = json.loads(temp_dic[keys])
                        temp_dic[keys] = js
                    except:
                        pass
            for key in new_data_dic:
                temp_dic[key] = new_data_dic[key]
        files = [testDataGenerator.path]
        for file in testDataGenerator.writers:
            files.append(file)
        size = 0
        for file in files:
            size += os.stat(file).st_size
        if size > 1000000:
            logger.debug("Source files are updating in a thread.")
            t = Thread(target=self.update_sources, args=(testDataGenerator,))
            t.daemon = True
            t.start()
            t.join()
        else:
            logger.debug("Source files are updating in main thread.")
            self.update_sources(testDataGenerator)

    def update_sources(self, testDataGenerator):
        if testDataGenerator.isUsecount:
                testDataGenerator.writer.save()  # saving source input file once everything is done
        for writer in testDataGenerator.writers:
            testDataGenerator.writers[writer].save()
            logger.debug(f"File updated: {writer}")

    def get_data_from_tdg(self, string, testDataGenerator):
        data = testDataGenerator.data_generators(string)
        if testDataGenerator.usecount_dict[repr(data[0])]["limit"]:
            data = [d for d in data if testDataGenerator.usecount_dict[repr(d)]["use"
                                                                ] < testDataGenerator.usecount_dict[repr(d)]["limit"]]
        if len(data) > 1:
            data = data[randint(0, len(data) - 1)]
        elif len(data) == 1:
            data = data[0]
        else:
            raise BaseException(f"Not enough data for {string}, please verify if data is present or usecount limit" \
                                "has reached!!")
        return data

    def rlp_process(self, string, fileName):
        # Will get real data from rlp_ prefix string
        rlp_string = self.__process_rlp_string(string)[5:-1]
        rlp_data = self.__rlp_string_to_python(rlp_string, fileName)
        data = rlp_data
        self.rlp_iterate(data, fileName)
        return data

    def rlp_iterate(self, data, fileName):
        # Rlp datas are stored in either json or list. This function will loop on every data and convert every
        # Rlp string to data
        if type(data) is list:
            for dt in data:
                if type(dt) is str:
                    dt = self.rlp_iterate(dt, fileName)
                elif type(dt) is dict or type(dt) is list:
                    dt = self.rlp_iterate(dt, fileName)
        elif type(data) is dict:
            for key in data:
                if type(data[key]) is list or type(data[key]) is dict:
                    data[key] = self.rlp_iterate(data[key], fileName)
                elif type(data[key]) is str:
                    data[key] = self.rlp_iterate(data[key], fileName)
        elif type(data) is str:
            if data[:4] == "RLP_":
                data = self.rlp_process(data, fileName)
        return data

    def __rlp_string_to_python(self, raw_data, fileName):
        # will convert rlp string to python
        sheetName = raw_data.split(',')[0].strip()
        headerName = raw_data.split(',')[1].strip().split('=')[0].strip()
        headerValue = raw_data.split(',')[1].strip().split('=')[1].strip()
        all_sheets, main_sheet = TestDataGenerator.read_excel(path=fileName, sheet_name=sheetName)
        data_list = []
        for data in main_sheet:
            main_value = data[headerName]
            if type(main_value) == float and str(main_value)[-2:] == '.0':
                main_value = str(int(main_value))
            if main_value.strip() == headerValue:
                for key in data:
                    try:
                        js = json.loads(data[key])
                        data[key] = js
                    except:
                        pass
                data_list.append(data)
        return data_list

    def __process_rlp_string(self, rlp_string):
        processed_string = ','.join([word.strip() for word in rlp_string.split(', ')])
        match = re.match(
            r"(RLP_(\(|\[)).+,.+=.+(\]|\))",
            processed_string
        )
        err_string = f"{rlp_string} not matching pattern RLP_(sheetName,HeaderName=DataToMatch"
        assert match, err_string
        return processed_string

    def __compareEqualStageInGlobalsAndDataRecord(self, currentNewRecordDict:dict) -> bool:
        """
        As method name says, compares, whether Stage in Global-settings is equal to stage in Data Record,
        so that this record might be excluded, if it's for the wrong Stage.

        :param currentNewRecordDict: The current Record
        :return: Boolean
        """
        lAppend = True
        if self.globals.get(GC.EXECUTION_STAGE):
            if currentNewRecordDict.get(GC.EXECUTION_STAGE):
                if currentNewRecordDict[GC.EXECUTION_STAGE] != self.globals[GC.EXECUTION_STAGE]:
                    lAppend = False
        return lAppend

    def readNextRecord(self):
        """
        We built self.range during init. Now we need to iterate over the range(s) in range,
        find appropriate record and return that - one at a time

        @return:
        """
        if len(self.rangeDict) == 0:
            # All records were processed
            return None
        try:
            # the topmost record of the RangeDict (RangeDict was built by the range(s) from the TestRun
            # - 1 because there's a header line in the Excel-Sheet.
            lRecord = self.dataDict[(list(self.rangeDict.keys())[0])]
            while not self.__compareEqualStageInGlobalsAndDataRecord(lRecord):
                logger.debug(f"Skipped record {str(lRecord)[:30]} due to wrong stage: {lRecord[GC.EXECUTION_STAGE]} vs. "
                             f"{self.globals[GC.EXECUTION_STAGE]}")
                self.rangeDict.pop(list(self.rangeDict.keys())[0])
                lRecord = self.dataDict[(list(self.rangeDict.keys())[0])]
        except Exception as e:
            logger.debug(f"Couldn't read record from database: {list(self.rangeDict.keys())[0]}")
            self.rangeDict.pop(list(self.rangeDict.keys())[0])
            return None

        # Remove the topmost entry from the rangeDict, so that next time we read the next entry in the lines above
        self.rangeDict.pop(list(self.rangeDict.keys())[0])
        return self.updateGlobals(lRecord)

    def readTestRecord(self, lineNumber=None):
        if lineNumber:
            self.lineNumber = lineNumber -1  # Base 0 vs. Base 1
        else:
            self.lineNumber += 1 # add 1 to read next line number

        try:
            record = self.dataDict[self.lineNumber]
            logger.info(f"Starting with Testrecord {self.lineNumber}, Details: " +
                        str({k: record[k] for k in list(record)[0:5]}))
            return self.updateGlobals(record)
        except Exception as e:
            logger.critical(f"Couldn't read record# {self.lineNumber}")

    def updateGlobals(self, record):

        self.globals[CGC.CUST_TOASTS] = ""
        self.globals[GC.TESTCASEERRORLOG] = ""
        self.globals[GC.TIMING_DURATION] = ""
        self.globals[GC.TIMELOG] = ""
        self.globals[GC.TESTCASESTATUS] = ""
        self.globals[GC.SCREENSHOTS] = ""

        record.update(self.globals)
        return record
